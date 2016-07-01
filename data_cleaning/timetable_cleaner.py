# Author Conor O'Kelly

# Aim of this script will be to take in the excel file data for the timetable. The data is first cleaned then
# placed into dicts that are returned to be handling by the main controller

import unittest
import openpyxl
import nose2


def phrase_excel_sheet_into_array_of_dicts(target_directory):

    # Function returns tuple of 0 is workbook object / 1 sheet objects
    sheet_object = load_workbook_return_sheets_object(target_directory)
    unmerge_excel_cells_and_perserve_data(sheet_object[1])

    sheet_object[0].save("1.test_data/new.xlsx")
    # Convert sheets into dict of all informaiton
    timetable_info = convert_sheets_into_dict(sheet_object[1])

    return timetable_info


def load_workbook_return_sheets_object(directory):

    try:
        # Load workbook with all sheets
        work_book = openpyxl.load_workbook(directory)

        # Load a all sheets from workbook
        sheet_names = work_book.get_sheet_names()

        # Remove last sheet name if name == All
        if sheet_names[-1] == "All" or sheet_names[-1] == "all":
            sheet_names.pop(-1)

        # Load sheet object into array
        sheet_objects = []
        for sh_name in sheet_names:
            sheet_objects.append(work_book.get_sheet_by_name(sh_name))

        return (work_book,sheet_objects)

    except FileNotFoundError:
        print("The file at", directory,"could not be found.")
        return (0,0)


def unmerge_excel_cells_and_perserve_data(sheet_objects):

    # Cycle through each cell range unmerge and replace data
    for current_sheet in sheet_objects:

        cell_range_in_sheet = [item for item in current_sheet.merged_cell_ranges]
        for cell_range in cell_range_in_sheet:
            # Unmerge range
            current_sheet.unmerge_cells(cell_range)

            # Set range coordinates and get first in merged range value
            first_coord, second_cord = cell_range.split(":")
            first_cell_value = current_sheet[first_coord].value

            # Generate cells id in range
            cells_in_merged_range = []
            for letter in range(ord(first_coord[0]),ord(second_cord[0])+1):
                letter_value = chr(letter)
                for number in range(int(first_coord[1:]),int(second_cord[1:])+1):
                    cells_in_merged_range.append(letter_value+str(number))

            for cell in cells_in_merged_range:
                current_sheet[cell].value = first_cell_value

    return sheet_objects


# Could come back and write this function to be able to handle a more dynamic approach #
def convert_sheets_into_dict(sheet_objects):

    timetable_information = []

    # Cycle through sheets and produce a dict for each one with timetable information
    for current_sheet in sheet_objects:
        current_sheet_array = []

        # Determine number of weeks on page by checking the first row for data
        if current_sheet["M2"].value is not None and current_sheet["B2"].value is not None:
            no_weeks = 2
        elif current_sheet["B2"].value is not None:
            no_weeks = 1
        else:
            no_weeks = 0

        for i in range(1,no_weeks+1):
            # Set starting point for week
            if i == 1:
                start_column = "B"
                finish_column = "J"
            elif i == 2:
                start_column = "N"
                finish_column = "V"
            else:
                start_column = "B"
                finish_column = "B"

            # Get month from top second column in each week timetable
            month = current_sheet[start_column+"1"].value.split(" ")[1]
            # Cell color

            # Cycle through columns and print first value
            for col_letter in range(ord(start_column),ord(finish_column)+1,2):
                day = current_sheet[chr(col_letter)+"2"].value
                date = day.strip() + " " + month
                # print(date)
                # Cycle through rows and extract the required data
                for row_no in range(3,11+1):
                    # print(chr(col_letter+1)+str(row_no))
                    room = current_sheet[chr(ord(start_column)-1)+"1"].value
                    time = current_sheet[chr(ord(start_column)-1)+str(row_no)].value
                    module = current_sheet[chr(col_letter)+str(row_no)].value
                    no_expected_students = current_sheet[chr(col_letter+1)+str(row_no)].value
                    cell_color = current_sheet[chr(col_letter)+str(row_no)].fill.start_color.index
                    # Check if cell has color and if has value indicating class did not go ahead
                    if cell_color != "00000000" and module is not None:
                        class_appeared_to_go_ahead = False
                    else:
                        class_appeared_to_go_ahead = True
                    time_slot ={"room": room, "date": date, "time": time, "module": module,
                                "no_expected_students": no_expected_students,
                                "class_appeared_to_go_ahead": class_appeared_to_go_ahead}

                    current_sheet_array.append(time_slot)

            # Add array for single week to file return information
            timetable_information.extend(current_sheet_array)

    # print(timetable_information)
    return timetable_information


# Unittesting

def test_workbook_loads_and_returns_workbook():
    results = load_workbook_return_sheets_object("1.test_data/B0.02 B0.03 B0.04 Timetable.xlsx")
    # Assert that first result of tuple is workbook objects
    assert isinstance(results[0], openpyxl.Workbook)


def test_unmerge_excel_cells_and_perserve_data():
    # Test the first return sheet has no merged cells
    assert True

if __name__ == '__main__':
    # x = phrase_excel_sheet_into_array_of_dicts("1.test_data/B0.02 B0.03 B0.04 Timetable.xlsx")
    nose2.main()