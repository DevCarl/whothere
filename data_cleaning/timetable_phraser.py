# Author Conor O'Kelly

# Aim of this script will be to take in the excel file data for the timetable. The data is first cleaned then
# placed into dicts that are returned to be handling by the main controller

# start_date is manually set at this point!!!!

import openpyxl
import copy
import nose2


def phrase_timetable_excel_sheet_into_array_of_dicts(target_directory):

    # Function returns tuple of 0 is workbook object / 1 sheet objects
    sheet_object = load_workbook_return_sheets_object(target_directory)
    unmerge_excel_cells_and_perserve_data(sheet_object[1])

    sheet_object[0].save(target_directory)
    # Convert sheets into dict of all informaiton
    timetable_info = convert_sheets_into_dict(sheet_object[1])

    # Cleaned module code up in data dicts
    cleaned_timetable_info = cleaned_moudle_data(timetable_info)

    return cleaned_timetable_info


def load_workbook_return_sheets_object(directory):

    try:
        # Load workbook with all sheets
        work_book = openpyxl.load_workbook(directory)

        # Load a all sheets from workbook
        sheet_names = work_book.get_sheet_names()

        # Remove last sheet name if name == All
        if sheet_names[-1] == "All" or sheet_names[-1] == "all":
            sheet_names.pop(-1)

        # Load sheet object into array
        sheet_objects = []
        for sh_name in sheet_names:
            sheet_objects.append(work_book.get_sheet_by_name(sh_name))

        return (work_book, sheet_objects)

    except FileNotFoundError:
        print("The file at", directory, "could not be found.")
        return (0,0)


def unmerge_excel_cells_and_perserve_data(sheet_objects):

    # Cycle through each cell range unmerge and replace data
    for current_sheet in sheet_objects:

        cell_range_in_sheet = [item for item in current_sheet.merged_cell_ranges]
        for cell_range in cell_range_in_sheet:
            # Unmerge range
            current_sheet.unmerge_cells(cell_range)

            # Set range coordinates and get first in merged range value
            first_coord, second_cord = cell_range.split(":")
            first_cell_value = current_sheet[first_coord].value

            # Generate cells id in range
            cells_in_merged_range = []
            for letter in range(ord(first_coord[0]),ord(second_cord[0])+1):
                letter_value = chr(letter)
                for number in range(int(first_coord[1:]),int(second_cord[1:])+1):
                    cells_in_merged_range.append(letter_value+str(number))

            for cell in cells_in_merged_range:
                current_sheet[cell].value = first_cell_value

    return sheet_objects


def cleaned_moudle_data(data_array):


    # Cycle through data dicts and fix any double modules / unknown module / career talks
    cleaned_data_array = []

    for dict in data_array:
        # First add news keys to all dicts
        dict["shared_time_slot"] = 0
        dict["practical"] = 0

        # cycle throught dicts and make corrections
        if dict.get("module") is not None and "&" in dict.get("module"):
            dict["shared_time_slot"] = 1
            # Create two copies of current dict
            first_module_dict = copy.deepcopy(dict)
            second_module_dict = copy.deepcopy(dict)

            # Separate out module code
            first_module_dict["module"] = dict.get("module").split("&")[0].strip()
            second_module_dict["module"] = dict.get("module").split("&")[1].strip()

            # Add new dict to cleaned data array
            cleaned_data_array.append(first_module_dict)
            cleaned_data_array.append(second_module_dict)

        elif dict.get("module") is not None and "(practical)" in dict.get("module").lower():
            dict["practical"] = 1
            dict["module"] = dict.get("module").replace("(Practical)", "").strip()
            # Add to cleaned data array
            cleaned_data_array.append(dict)

        elif dict.get("module") is not None and "(lecture)" in dict.get("module").lower():
            dict["module"] = dict.get("module").replace("(Lecture)", "").strip()
            # Add to cleaned data array
            cleaned_data_array.append(dict)

        elif dict.get("module") is not None and "Booked by School of CS (no other data available)" in dict.get("module"):
            dict["module"] = dict.get("module").replace("Booked by School of CS (no other data available)", "CS_school").strip()
            # Add to cleaned data array
            cleaned_data_array.append(dict)

        elif dict.get("module") is not None and "Career opportunities talks" in dict.get("module"):
            dict["module"] = dict.get("module").replace("Career opportunities talks", "Careers").strip()
            # Add to cleaned data array
            cleaned_data_array.append(dict)

        else:
            # Add to cleaned data array
            cleaned_data_array.append(dict)

    return cleaned_data_array

# Could come back and write this function to be able to handle a more dynamic approach #
def convert_sheets_into_dict(sheet_objects):

    timetable_information = []

    # Cycle through sheets and produce a dict for each one with timetable information
    for current_sheet in sheet_objects:
        # print(current_sheet)
        current_sheet_array = []

        # Determine number of weeks on page by checking the first row for data
        if current_sheet["M2"].value is not None and current_sheet["B2"].value is not None:
            no_weeks = 2
        elif current_sheet["B2"].value is not None:
            no_weeks = 1
        else:
            no_weeks = 0

        for i in range(1, no_weeks+1):
            # print(i,"week")
            # Set starting point for week
            if i == 1:
                start_column = "B"
                finish_column = "J"
            elif i == 2:
                start_column = "N"
                finish_column = "V"
            else:
                start_column = "B"
                finish_column = "B"

            # Get month from top second column in each week timetable
            month = current_sheet[start_column+"1"].value.split(" ")[1]
            # Cell color

            ### Section need to be updated ###

            # Set start date manually
            if i == 1:
                start_day = 2
            else:
                start_day = 9

            # Cycle through columns and print first value
            for col_letter in range(ord(start_column),ord(finish_column)+1,2):

                ### Section need to be updated ###
                # print(col_letter)
                date = str(start_day) + "/11/15"
                start_day += 1
                # print(date)

                # Cycle through rows and extract the required data
                for row_no in range(3, 11+1):
                    room = current_sheet[chr(ord(start_column)-1)+"1"].value.replace(".", "")
                    time = current_sheet[chr(ord(start_column)-1)+str(row_no)].value
                    module = current_sheet[chr(col_letter)+str(row_no)].value
                    no_expected_students = current_sheet[chr(col_letter+1)+str(row_no)].value
                    cell_color = current_sheet[chr(col_letter)+str(row_no)].fill.start_color.index

                    # Check if cell has color and if has value indicating class did not go ahead
                    if cell_color != "00000000" and module is not None:
                        class_appeared_to_go_ahead = False
                    else:
                        class_appeared_to_go_ahead = True
                    time_slot ={"room": room, "date": date, "time": time, "module": module,
                                "no_expected_students": no_expected_students,
                                "class_appeared_to_go_ahead": class_appeared_to_go_ahead}

                    current_sheet_array.append(time_slot)

        # Add array for single week to file return information
        timetable_information.extend(current_sheet_array)

    # print(len(timetable_information))
    # count = 1
    # for i in timetable_information:
    #     print(i, count)
    #     count += 1
    return timetable_information


# Nose2 tests

# def test_workbook_loads_and_returns_workbook():
#     results = load_workbook_return_sheets_object("1.test_data/B0.02 B0.03 B0.04 Timetable.xlsx")
#     # Assert that first result of tuple is workbook objects
#     assert isinstance(results[0], openpyxl.Workbook)
#
#
# def test_unmerge_excel_cells_and_perserve_data():
#     # Test the first return sheet has no merged cells
    assert True

if __name__ == '__main__':
    x = phrase_timetable_excel_sheet_into_array_of_dicts("1.test_data/B0.02 B0.03 B0.04 Timetable.xlsx")
    print(x)
    # nose2.main()